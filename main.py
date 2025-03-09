# 📌 Importamos las librerías necesarias
from openpyxl import load_workbook
import pandas as pd
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Border, Side
from openpyxl.chart import BarChart, PieChart, Reference

# 📌 Cargar el archivo Excel y activar la hoja principal
data = load_workbook("data/Financial Sample.xlsx")
sheet = data.active
sheet.title = "Datos originales"

# 📌 Crear y activar la hoja de "Resumenes" para almacenar los cálculos
data.create_sheet("Resumenes")
data.active = data["Resumenes"]
sheet = data.active

# 📌 Cargar los datos con Pandas
df = pd.read_excel("data/Financial Sample.xlsx", engine="openpyxl")

# 📌 Cálculo de estadísticas financieras por país

ventas_brutas_por_pais = df.groupby("Country")["Gross Sales"].sum().round(0)
ventas_brutas_por_pais.name = "Ventas Brutas (€)"

promedio_ventas = df.groupby("Country")["Gross Sales"].mean().round(0)
promedio_ventas.name = "Promedio Ventas (€)"

total_ganancias = df.groupby("Country")["Profit"].sum().round(0)
total_ganancias.name = "Ganancias totales (€)"

df["Profit_margin"] = df["Profit"] / df["Gross Sales"]
margen_ganancia_medio_por_pais = df.groupby("Country")["Profit_margin"].mean().round(6) * 100
margen_ganancia_medio_por_pais.name = "Margen de ganancia medio (%)"

promedio_unidades_vendidas = df.groupby("Country")["Units Sold"].mean().round(2)
promedio_unidades_vendidas.name = "Promedio Unidades vendidas (ud)"

df['Ingresos'] = df['Sale Price'] * df['Units Sold']
ingresos_totales_por_pais = df.groupby('Country')['Ingresos'].sum().round(0)
ingresos_totales_por_pais.name = "Ingresos totales (€)"

df['Year Growth'] = df.groupby('Country')['Gross Sales'].pct_change()
crecimiento_anual_por_pais = df.groupby('Country')['Year Growth'].mean().round(2)
crecimiento_anual_por_pais.name = "Crecimiento anual (%)"

# 📌 Almacenar estadísticas en la hoja "Resumenes"
paises = ["Estadísticas", "Canada", "Francia", "Alemania", "Mexico", "USA"]
estadisticas = {
    ventas_brutas_por_pais.name: ventas_brutas_por_pais,
    promedio_ventas.name: promedio_ventas,
    ingresos_totales_por_pais.name: ingresos_totales_por_pais,
    total_ganancias.name: total_ganancias,
    promedio_unidades_vendidas.name: promedio_unidades_vendidas,
    margen_ganancia_medio_por_pais.name: margen_ganancia_medio_por_pais,
    crecimiento_anual_por_pais.name: crecimiento_anual_por_pais
}

sheet.append(paises)  # Agregamos la fila de encabezados

# Insertamos los valores de cada estadística
for nombre, stat in estadisticas.items():
    sheet.append([nombre] + [round(dato, 2) for dato in stat])

# 📌 Ajustar el ancho de las columnas automáticamente
for col in sheet.columns:
    max_length = max((len(str(cell.value)) for cell in col if cell.value), default=0)
    col_letter = col[0].column_letter
    sheet.column_dimensions[col_letter].width = max_length + 2

# 📌 Aplicar formato condicional (escala de colores)
for row in sheet.iter_rows(min_row=2, min_col=2):
    valores = [int(cell.value) for cell in row]
    color_scale = ColorScaleRule(start_type="num", start_value=min(valores), start_color="f80a0a",
                                 mid_type="num", mid_value=sum(valores) / len(valores), mid_color="fcf80d",
                                 end_type="num", end_value=max(valores), end_color="7af80a")
    sheet.conditional_formatting.add(f"{row[0].coordinate}:{row[-1].coordinate}", color_scale)

# 📌 Agregar bordes a la tabla
border_style = Border(left=Side(style="thin"), right=Side(style="thin"),
                      top=Side(style="thin"), bottom=Side(style="thin"))

for row in sheet.iter_rows(min_row=1, max_row=8, min_col=1, max_col=6):
    for cell in row:
        cell.border = border_style

# 📌 Crear la hoja "Graficos" para los gráficos
data.create_sheet("Graficos")
data.active = data["Graficos"]
sheet = data.active

## 📊 Creación de gráficos ##

# 📌 Gráfico de Barras: Ventas Brutas y Ganancias Totales por país
bar_chart = BarChart()
paises_ref = Reference(data["Resumenes"], min_col=2, min_row=1, max_col=6)
gross_data = Reference(data["Resumenes"], min_col=1, min_row=2, max_col=6)
profit_data = Reference(data["Resumenes"], min_col=1, min_row=4, max_col=6)

bar_chart.add_data(gross_data, from_rows=True, titles_from_data=True)
bar_chart.add_data(profit_data, from_rows=True, titles_from_data=True)
bar_chart.set_categories(paises_ref)
bar_chart.title = "Comparación de Ventas y Ganancias"
bar_chart.y_axis.title = "€"
bar_chart.x_axis.title = "Países"
bar_chart.style = 18
sheet.add_chart(bar_chart, "C3")

# 📌 Gráfico de Barras: Promedio de Ventas por País
bar_chart = BarChart()
promedio_ventas_data = Reference(data["Resumenes"], min_col=1, min_row=3, max_col=6)
bar_chart.add_data(promedio_ventas_data, from_rows=True, titles_from_data=True)
bar_chart.set_categories(paises_ref)
bar_chart.title = "Promedio de ventas por país"
sheet.add_chart(bar_chart, "C20")

# 📌 Gráfico Circular: Crecimiento Anual por País
pie_chart = PieChart()
data_crecimiento = Reference(data["Resumenes"], min_col=1, min_row=8, max_col=6)
pie_chart.add_data(data_crecimiento, from_rows=True, titles_from_data=True)
pie_chart.title = "Crecimiento anual por país"
sheet.add_chart(pie_chart, "M20")

# 📌 Guardar el archivo Excel
data.save("data/Financial Sample.xlsx")

print("✅ Proceso completado con éxito")
